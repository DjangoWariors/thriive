"""ReportResult → branded XLSX bytes (openpyxl)."""
import io
from decimal import Decimal

from .base import BRAND_PRIMARY, BRAND_TINT, ReportResult

# Indian grouping for money/quantity; percent gets its own format.
_NUMBER_FORMATS = {
    'decimal': '##,##,##0.00',
    'integer': '##,##,##0',
    'percent': '0.00"%"',
}


def render(result: ReportResult) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'

    header_fill = PatternFill('solid', fgColor=BRAND_PRIMARY)
    header_font = Font(bold=True, color='FFFFFF')
    tint_fill = PatternFill('solid', fgColor=BRAND_TINT)
    title_font = Font(bold=True, size=14, color=BRAND_PRIMARY)

    ncols = len(result.columns)

    # Title band.
    ws.cell(row=1, column=1, value=result.title).font = title_font
    if result.confidential:
        c = ws.cell(row=1, column=ncols if ncols else 1, value='CONFIDENTIAL')
        c.font = Font(bold=True, color=BRAND_PRIMARY)
        c.alignment = Alignment(horizontal='right')

    header_row = 3
    for ci, col in enumerate(result.columns, start=1):
        cell = ws.cell(row=header_row, column=ci, value=col.label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows.
    r = header_row
    for ri, row in enumerate(result.rows, start=header_row + 1):
        r = ri
        for ci, col in enumerate(result.columns, start=1):
            value = row.get(col.key)
            cell = ws.cell(row=ri, column=ci, value=_coerce(value, col.type))
            fmt = _NUMBER_FORMATS.get(col.type)
            if fmt:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal='right')
            if ri % 2 == 0:
                cell.fill = tint_fill

    # Totals row.
    if result.summary:
        r += 1
        for ci, col in enumerate(result.columns, start=1):
            value = result.summary.get(col.key)
            label = 'TOTAL' if ci == 1 and value is None else value
            cell = ws.cell(row=r, column=ci, value=_coerce(label, col.type))
            cell.font = Font(bold=True)
            fmt = _NUMBER_FORMATS.get(col.type)
            if fmt and value is not None:
                cell.number_format = fmt

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Column widths.
    for ci, col in enumerate(result.columns, start=1):
        width = col.width or max(len(col.label) + 2, 12)
        ws.column_dimensions[get_column_letter(ci)].width = min(width, 50)

    _info_sheet(wb, result)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _coerce(value, col_type: str):
    if value is None:
        return None
    if col_type in ('decimal', 'percent'):
        return float(value) if isinstance(value, (Decimal, int, float, str)) and _num(value) else value
    if col_type == 'integer':
        return int(value) if _num(value) else value
    return str(value)


def _num(value) -> bool:
    try:
        float(value)
        return True
    except (TypeError, ValueError):
        return False


def _info_sheet(wb, result: ReportResult):
    ws = wb.create_sheet('Report Info')
    ws['A1'] = 'Field'
    ws['B1'] = 'Value'
    rows = [
        ('Title', result.title),
        ('Rows', result.row_count),
        ('Confidential', 'Yes' if result.confidential else 'No'),
    ]
    for k, v in (result.meta or {}).items():
        rows.append((k.replace('_', ' ').title(), str(v)))
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 60
